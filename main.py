import tkinter as tk
from tkinter import filedialog
import openpyxl

def read_excel_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    dictionary = {}
    
    found_integer = False
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[0]
        value_c = row[2]
        value_e = row[4]
        
        if not found_integer:
            if isinstance(key, int):
                found_integer = True
                if key != '' and value_c != '' and value_e != '':
                    dictionary[key] = {'descrizione': value_c, 'prezzo': value_e, 'nomefile': file_path}
        else:
            if key != '':
                dictionary[key] = {'descrizione': value_c, 'prezzo': value_e, 'nomefile': file_path}
    
    return dictionary

def merge_and_filter_dictionaries(dict1, dict2):
    common_keys = set(dict1.keys()) & set(dict2.keys())
    merged_dict = {}
    
    for key in common_keys:
        price1 = dict1[key]['prezzo'] if dict1.get(key) else None
        price2 = dict2[key]['prezzo'] if dict2.get(key) else None
        
        if price1 is not None and price2 is not None:
            if price1 < price2:
                merged_dict[key] = dict1[key]
            else:
                merged_dict[key] = dict2[key]
    
    return merged_dict

def browse_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def process_files():
    file_path1 = entry_file_path1.get()
    file_path2 = entry_file_path2.get()
    
    if file_path1 and file_path2:
        dict1 = read_excel_file(file_path1)
        dict2 = read_excel_file(file_path2)
        
        merged_dict = merge_and_filter_dictionaries(dict1, dict2)
        
        # Creazione di un nuovo file Excel
        new_file = openpyxl.Workbook()
        sheet = new_file.active
        
        # Scrittura dei dati nel foglio di lavoro
        sheet['A1'] = 'Chiave'
        sheet['B1'] = 'Descrizione'
        sheet['C1'] = 'Prezzo'
        sheet['D1'] = 'Nome File'
        
        row = 2
        for key, value in merged_dict.items():
            sheet.cell(row=row, column=1).value = key
            sheet.cell(row=row, column=2).value = value['descrizione']
            sheet.cell(row=row, column=3).value = value['prezzo']
            sheet.cell(row=row, column=4).value = value['nomefile']
            row += 1
        
        # Salvataggio del file Excel
        new_file.save('merged_data.xlsx')
        new_file.close()
        
        label_message.config(text="File Excel salvato correttamente.")
    else:
        label_message.config(text="Seleziona entrambi i file.")


# Creazione dell'interfaccia grafica utilizzando Tkinter
window = tk.Tk()
window.title("Merge Excel Files")
window.geometry("400x200")

# Label e Entry per il percorso del file 1
label_file_path1 = tk.Label(window, text="Percorso del file 1:")
label_file_path1.pack()
entry_file_path1 = tk.Entry(window)
entry_file_path1.pack()

# Pulsante di navigazione per il file 1
button_browse1 = tk.Button(window, text="Sfoglia", command=lambda: browse_file(entry_file_path1))
button_browse1.pack()

# Label e Entry per il percorso del file 2
label_file_path2 = tk.Label(window, text="Percorso del file 2:")
label_file_path2.pack()
entry_file_path2 = tk.Entry(window)
entry_file_path2.pack()

# Pulsante di navigazione per il file 2
button_browse2 = tk.Button(window, text="Sfoglia", command=lambda: browse_file(entry_file_path2))
button_browse2.pack()

# Pulsante per avviare il processo di unione e filtraggio
button_process = tk.Button(window, text="Confronta", command=process_files)
button_process.pack()

label_message = tk.Label(window)
label_message.pack()

window.mainloop()
